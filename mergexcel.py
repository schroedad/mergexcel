#!/usr/bin/env python3

import os
import sys
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Terminal styling
def clear_line():
    sys.stdout.write('\033[K')
    sys.stdout.flush()

def progress_bar(current, total, message):
    bar_length = 20
    filled_length = int(round(bar_length * current / total))
    bar = '█' * filled_length + '-' * (bar_length - filled_length)

    # Clear the entire line and reset cursor to start
    print(f'\r\033[K|{bar}| {message}', end='', flush=True)

# Main function
def merge_excel():
    print("\n📚 Welcome to mergexcel!\n")
    print("This will merge all Excel (.xlsx) files from a specified directory into a single combined Excel file.")

    folder = input("\nSpecify the folder containing excel files [./]: ") or "./"
    folder = os.path.expanduser(folder)
    codebook_count = input("How many codebook worksheets are there [1]? ") or "1"

    try:
        codebook_count = int(codebook_count)
    except ValueError:
        print("❌ Invalid number for codebooks.")
        return

    default_output = os.path.join(os.getcwd(), 'combined_excel.xlsx')
    output_file = input(f"Where should the combined file be saved [./combined_excel.xlsx]? ") or default_output

    print()

    excel_files = sorted(glob.glob(os.path.join(folder, '*.xlsx')))

    if not excel_files:
        print("❌ No Excel files found.")
        return

    target_wb = Workbook()
    first_file = excel_files.pop(0)
    source_wb = load_workbook(first_file)
    worksheets = source_wb.sheetnames

    # Setup sheets in target workbook
    for idx, ws_name in enumerate(worksheets):
        source_ws = source_wb[ws_name]
        target_ws = target_wb.active if idx == 0 else target_wb.create_sheet(title=ws_name)
        for row in source_ws.iter_rows(values_only=True):
            target_ws.append(row)

    total_files = len(excel_files)
    current_file = 1
    errors = False
    summary_mismatch = {}

    expected_sheets = worksheets[codebook_count:]

    for file in excel_files:
        progress_bar(current_file, total_files + 2, f"Opening {os.path.basename(file)}")
        wb = load_workbook(file)
        actual_sheets = wb.sheetnames[codebook_count:] if codebook_count else wb.sheetnames

        matched_sheets = set(actual_sheets).intersection(expected_sheets)

        missing_count = len(expected_sheets) - len(matched_sheets)
        extra_count = len(actual_sheets) - len(matched_sheets)

        if missing_count or extra_count:
            summary_mismatch[os.path.basename(file)] = {'missing': missing_count, 'extra': extra_count}

        if not matched_sheets:
            print(f"\n❌ No matching worksheets in {file}. Stopping.")
            errors = True
            break

        for ws_name in matched_sheets:
            progress_bar(current_file, total_files + 2, f"Copying {ws_name} from {os.path.basename(file)}")
            ws = wb[ws_name]
            target_ws = target_wb[ws_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                target_ws.append(row)

        current_file += 1

    if not errors:
        # Apply formatting
        progress_bar(total_files + 1, total_files + 2, "Applying final formatting")
        for ws_name in expected_sheets:
            ws = target_wb[ws_name]
            header_row = next(ws.iter_rows(max_row=1))
            for cell in header_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

            # Adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    target_wb.save(output_file)

    if errors:
        print(f"\n⚠️ Stopped due to errors. Partial combined file saved as '{output_file}'.")
    else:
        progress_bar(total_files + 2, total_files + 2, "Done! ✅")
        print(f"\n\n🎉 Successfully merged Excel files into '{output_file}'.")

    print("\nSummary:")
    print(f"Total files merged: {current_file}")
    print(f"Total worksheets processed: {len(worksheets)}")

    if summary_mismatch:
        print("\nWorksheet mismatches:")
        for fname, counts in summary_mismatch.items():
            print(f"- Missing: {counts['missing']:2} | Extra: {counts['extra']:2} | {fname}")

if __name__ == "__main__":
    merge_excel()